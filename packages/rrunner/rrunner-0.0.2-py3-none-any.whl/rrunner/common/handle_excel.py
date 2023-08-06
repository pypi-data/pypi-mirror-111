import openpyxl


# 定义一个类从excel中读数据和写数据，需要安装openpyxl插件
class Excel:
    def __init__(self, file_name, sheet_name) -> object:
        # 定义一个excel文件夹名字
        self.file_name = file_name
        # 定义一个excel中工作薄名字
        self.sheet_name = sheet_name

    # 从excel中读数据，读出来的数据处理成list的形式
    def read_excel(self):
        # 获取工作簿对象
        wb = openpyxl.load_workbook(self.file_name)
        # 选择表单
        sh = wb[self.sheet_name]
        # 按行获取表单所有格子中的数据，每一行的数据放在一个元组中
        res = list(sh.rows)
        # 获取第一行的数据，作为字典的键
        title = [item.value for item in res[0]]
        # 创建一个空列表，用来存放用例数据
        data_cases = []
        # 遍历除第一行之外的数据
        for item in res[1:]:
            # 获取该行数据的值
            data = [c.value for c in item]
            # 将该行数据和title（第一行数据）打包转换为字典
            data_case = dict(zip(title, data))
            # 将转换的字典添加到前面创建的空列表data_cases中
            data_cases.append(data_case)
        return data_cases

    # 往excel中某行某列中写数据
    def write_excel(self, row, column, value):
        """写入数据"""
        wb = openpyxl.load_workbook(self.file_name)
        sh = wb[self.sheet_name]
        # 写入数据
        sh.cell(row=row, column=column, value=value)
        # 保存文件
        wb.save(self.file_name)
