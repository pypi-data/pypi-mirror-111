# -*- coding:utf-8 -*-

import requests
import os
import openpyxl

from rrunner.common.handle_config import config
from rrunner.common.handle_path import DATA_DIR, CASE_DIR


def getByPath(path, obj):
    paths = path.split(".")
    for path in paths:
        obj = obj.get(path, None)
        if obj == None:
            break
    return obj


def dic2String(obj):
    rs = ['{']
    for kv in obj.items():
        rs.append('\"' + kv[0] + "\":\"" + str(kv[1]) + '\"')
        rs.append(',')
    if len(rs) > 2:
        rs.pop()
    rs.append('}')
    return ''.join(rs)


def parseBody(parItem, body, raw):
    if parItem.get('schema', None) != None:
        refPath = getByPath('schema.$ref', parItem)
        if refPath == None:
            # 数组
            refPath = getByPath('schema.items.$ref', parItem)

        if refPath != None:
            refPath = refPath.replace('#/definitions/', '')

            refData = getByPath('definitions.' + refPath + '.properties', raw)

            if refData != None:
                for ri in refData.items():
                    body[ri[0]] = (0 if ri[1].get('type', None) == 'integer'
                                   else "")
            elif parItem.get('description', None) != None:
                body['_parms'] = parItem['description']
            else:
                body[parItem['name']] = ''
    else:
        body[parItem['name']] = (0 if parItem.get('type', None) == 'integer'
                                 else "")


def writeRow(func, ws, i):
    i = str(i)
    ws['A' + i] = func['case_id']
    ws['B' + i] = func['title']
    ws['C' + i] = func['interface']
    ws['D' + i] = func['content-type']
    ws['E' + i] = func['method']
    ws['F' + i] = func['url']
    ws['G' + i] = func['data']
    ws['H' + i] = func['expected']
    ws['I' + i] = func['check_sql']
    ws['J' + i] = func['result']
    ws['K' + i] = func['tag']


def writeCaseClass(cName):
    caseName = 'test_' + cName + '_controller.py'
    dataName = 'test_' + cName + '_controller.xlsx'
    isExist = os.path.exists(os.path.join(CASE_DIR + "\InnerApi", caseName))

    if isExist:
        return
    f = open(os.path.join(CASE_DIR + "\InnerApi", caseName), 'w')
    f.write("import os\n")
    f.write("import allure\n")
    f.write("import pytest\n")
    f.write("from common.handle_excel import Excel\n")
    f.write("from common.handle_path import DATA_DIR\n")
    f.write("from common.handle_config import config\n")
    f.write("from common.requtest_assert import RequestsAssert\n")
    f.write("class Test" + cName + ":\n")
    f.write('    excel = Excel(os.path.join(DATA_DIR, "{}"), "Sheet")\n'.format(dataName))
    f.write("    test_data = excel.read_excel()\n")
    f.write('    module = config.get("test_data", "module")\n')
    f.write('    if module == "0":\n')
    f.write('        for i in range(0, len(test_data) - 1):\n')
    f.write('            if None == test_data[i]["tag"]:\n')
    f.write('                del (test_data[i])\n')
    f.write('    @allure.feature("{}")\n'.format(cName))
    f.write("    @pytest.mark.parametrize('item', test_data)\n")
    f.write('    def test_' + cName + '(self, item, get_token):\n')
    f.write("        headers = get_token\n")
    f.write("        res = RequestsAssert.apiRequest(item, headers)\n")
    f.write("        write = self.excel.write_excel\n")
    f.write("        RequestsAssert.apiAssert(res, item, write)\n")


def writeCase(cName, funcs):
    caseName = 'test_' + cName + '_controller.xlsx'
    isExist = os.path.exists(os.path.join(DATA_DIR, caseName))

    if isExist:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    i = 1
    for func in funcs:
        writeRow(func, ws, i)
        i += 1

    wb.save(os.path.join(DATA_DIR, caseName))


def main(catName, rules):
    rs = requests.get(config.get("env", "swagger_url"))
    raw = rs.json()
    paths = getByPath("paths", raw)
    funcs = []
    lastCName = None
    i = 1
    keys = paths.keys()
    # keys.sort()
    keys = sorted(keys)

    for pKey in keys:
        path = pKey
        value = paths[pKey]
        cName = path.split('/')[1]

        if catName != '*' and cName != catName:
            continue

        if lastCName != cName and lastCName != None:
            writeCase(lastCName, funcs)
            writeCaseClass(lastCName)
            i = 1
            funcs = []
        lastCName = cName

        method = 'post' if value.get('post', None) != None else 'get'

        value = value[method]
        params = getByPath("parameters", value)

        desc = getByPath("summary", value)

        body = {}
        query = {}
        data = {}

        for par in params:
            if par['in'] == 'body':
                parseBody(par, body, raw)
            elif par['in'] == 'query':
                query[par['name']] = ''

        data = {'query': query, 'body': body}
        # if len(body) > 0 and len(query) > 0:
        #     data = {query: query, body: body}
        # else:
        #     data = body if len(body) > 0 else query

        if i == 1:
            funcs.append({
                'case_id': 'case_id',
                'title': 'title',
                'content-type': 'content-type',
                'interface': 'interface',
                'url': 'url',
                'method': 'method',
                'data': 'data',
                'expected': 'expected',
                'check_sql': 'check_sql',
                'result': 'result',
                'tag': 'tag'
            })

        item = {
            'case_id': str(i),
            'title': desc,
            'content-type': 'union',
            'interface': path,
            'url': "/smartfactory" + path,
            'method': method,
            'data': '',
            'expected': '{\"innerCode\":"200"}',
            'check_sql': '',
            'result': '',
            'tag': ''
        }

        if len(body) > 0:
            item['content-type'] = 'data'
            if len(body) == 1 and body.get('_parms', None) != None:
                item['data'] = body['_parms']
            else:
                item['data'] = dic2String(body)
        else:
            item['content-type'] = 'params'
            item['data'] = dic2String(query)
        if method == "post":
            item['content-type'] = 'json'
        else:
            item['content-type'] = 'params'

        funcs.append(item)

        i += 1

    writeCase(lastCName, funcs)
    writeCaseClass(lastCName)


def parseArgs():
    args = {
        'int': {
            'min': 0,
            'max': 100
        },
        'string': {
            'min': 0,
            'max': 100,
            'whiteSpace': True,
            'required': True
        }
    }

    return args


main('*', parseArgs())
