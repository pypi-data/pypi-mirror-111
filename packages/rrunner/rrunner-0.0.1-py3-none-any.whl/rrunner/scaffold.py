import os.path
import subprocess
import sys

import openpyxl


from loguru import logger
from sentry_sdk import capture_message


def init_parser_scaffold(subparsers):
    sub_parser_scaffold = subparsers.add_parser(
        "startproject", help="Create a new project with template structure."
    )
    sub_parser_scaffold.add_argument(
        "project_name", type=str, nargs="?", help="Specify new project name."
    )
    return sub_parser_scaffold


def create_scaffold(project_name):
    """ create scaffold with specified project name.
    """

    def show_tree(prj_name):
        try:
            print(f"\n$ tree {prj_name} -a")
            subprocess.run(["tree", prj_name, "-a"])
            print("")
        except FileNotFoundError:
            logger.warning("tree command not exists, ignore.")

    if os.path.isdir(project_name):
        logger.warning(
            f"Project folder {project_name} exists, please specify a new project name."
        )
        show_tree(project_name)
        return 1
    elif os.path.isfile(project_name):
        logger.warning(
            f"Project name {project_name} conflicts with existed file, please specify a new one."
        )
        return 1

    logger.info(f"Create new project: {project_name}")
    print(f"Project Root Dir: {os.path.join(os.getcwd(), project_name)}\n")

    def create_folder(path):
        os.makedirs(path)
        msg = f"created folder: {path}"
        print(msg)

    def create_file(path, file_content=""):
        with open(path, "w", encoding="utf-8") as f:
            f.write(file_content)
        msg = f"created file: {path}"
        print(msg)

    content = """
import pytest
import requests
from jsonpath import jsonpath
from rrunner.common.handle_sql import db
# from tools.jt_role_handle import get_factory_all_user, remove_alluser_allrole, get_menu_id, get_user, add_role, \
#     random_name, add_device, add_place_id
from rrunner.common.handle_config import config, configtestdata
import time
from random import randint
from rrunner.common.requtest_assert import RequestsAssert
import random

condition = config.get("test_data", "condition")


# 登录返回headers值
@pytest.fixture(scope='session', autouse=True)
def get_token():
    url = config.get("env", "base_url") + config.get("test_data", "url")
    print(url)
    params = {"username": config.get("test_data", "user"), "password": config.get("test_data", "password")}
    print(params)
    headers = eval(config.get("env", "headers"))
    response = requests.post(url=url, json=params, headers=headers).json()
    print(response)
    token = jsonpath(response, "$..token")[0]
    headers["Authorization"] = token
    url = config.get("env", "base_url") + "/smartfactory/factory/join"
    print(url)
    params = {"id": config.get("test_data", "factory_id")}
    response = requests.post(url=url, params=params, headers=headers).json()
    print(response)
    token = jsonpath(response, "$..result")[0]
    headers["Authorization"] = token
    # 不新增工厂时可以不用运行
    # add_place_id()
    # test_device_id = add_device()
    yield headers
    if condition == "0":
        sql = "select * from smartfactory.smt_device_place where factory_id={}".format(
            config.get("test_data", "factory_id"))
        device_PlaceId = db.find_data(sql)[0]["id"]
        sql = "delete from smartfactory.smt_device_place where id={}".format(device_PlaceId)
        db.find_data(sql)
        for i in range(0, 8):
            # sql = "delete from smartfactory.smt_device where id={}".format(test_device_id[i]["id"])
            db.find_data(sql)
    """

    config_content = """[env]
base_url1 = http://120.42.34.82:83/api/vendor/leizhizao
base_url = http://smartfactoryback.letsbim.net
base_url_mes=http://47.110.146.39:8123
base_url_demo = http://smartfactoryback.letsbim.net
base_url_test = http://192.168.250.79:7081
base_urloa = http://smartfactoryback.letsbim.net
headers = {'User-Agent': 'Mozilla/5.0'}
swagger_url = http://192.168.250.78:7081/smartfactory/v2/api-docs?group=v1

[mysql]
host = 119.3.12.34
port = 3306
user = root
password = gemhone2016
password1 = P@ssw0rd
password_demo = gemhone2016
host_test = 192.168.250.79
host_demo = 119.3.12.34

[test_data]
url = /smartfactory/login/oa
user = zhuhongying
password = 123456
factory_id = 1300982115276120068
#配置module的值为0执行smoke用例，为1执行all用例
module = 1
#配置condition的值为0执行后执行后置条件，为1执行后不执行后置条件
condition=1

    """

    demo_testcase_request_content = """
config:
    name: "request methods testcase with functions"
    variables:
        foo1: config_bar1
        foo2: config_bar2
        expect_foo1: config_bar1
        expect_foo2: config_bar2
    base_url: "https://postman-echo.com"
    verify: False
    export: ["foo3"]

teststeps:
-
    name: get with params
    variables:
        foo1: bar11
        foo2: bar21
        sum_v: "${sum_two(1, 2)}"
    request:
        method: GET
        url: /get
        params:
            foo1: $foo1
            foo2: $foo2
            sum_v: $sum_v
        headers:
            User-Agent: HttpRunner/${get_httprunner_version()}
    extract:
        foo3: "body.args.foo2"
    validate:
        - eq: ["status_code", 200]
        - eq: ["body.args.foo1", "bar11"]
        - eq: ["body.args.sum_v", "3"]
        - eq: ["body.args.foo2", "bar21"]
-
    name: post raw text
    variables:
        foo1: "bar12"
        foo3: "bar32"
    request:
        method: POST
        url: /post
        headers:
            User-Agent: HttpRunner/${get_httprunner_version()}
            Content-Type: "text/plain"
        data: "This is expected to be sent back as part of response body: $foo1-$foo2-$foo3."
    validate:
        - eq: ["status_code", 200]
        - eq: ["body.data", "This is expected to be sent back as part of response body: bar12-$expect_foo2-bar32."]
-
    name: post form data
    variables:
        foo2: bar23
    request:
        method: POST
        url: /post
        headers:
            User-Agent: HttpRunner/${get_httprunner_version()}
            Content-Type: "application/x-www-form-urlencoded"
        data: "foo1=$foo1&foo2=$foo2&foo3=$foo3"
    validate:
        - eq: ["status_code", 200]
        - eq: ["body.form.foo1", "$expect_foo1"]
        - eq: ["body.form.foo2", "bar23"]
        - eq: ["body.form.foo3", "bar21"]
"""
    demo_testcase_with_ref_content = """
config:
    name: "request methods testcase: reference testcase"
    variables:
        foo1: testsuite_config_bar1
        expect_foo1: testsuite_config_bar1
        expect_foo2: config_bar2
    base_url: "https://postman-echo.com"
    verify: False

teststeps:
-
    name: request with functions
    variables:
        foo1: testcase_ref_bar1
        expect_foo1: testcase_ref_bar1
    testcase: testcases/demo_testcase_request.yml
    export:
        - foo3
-
    name: post form data
    variables:
        foo1: bar1
    request:
        method: POST
        url: /post
        headers:
            User-Agent: HttpRunner/${get_httprunner_version()}
            Content-Type: "application/x-www-form-urlencoded"
        data: "foo1=$foo1&foo2=$foo3"
    validate:
        - eq: ["status_code", 200]
        - eq: ["body.form.foo1", "bar1"]
        - eq: ["body.form.foo2", "bar21"]
"""
    ignore_content = "\n".join(
        [".env", "reports/*", "__pycache__/*", "*.pyc", ".python-version", "logs/*"]
    )
    demo_debugtalk_content = """import time

from httprunner import __version__


def get_httprunner_version():
    return __version__


def sum_two(m, n):
    return m + n


def sleep(n_secs):
    time.sleep(n_secs)
"""
    demo_env_content = "\n".join(["USERNAME=leolee", "PASSWORD=123456"])

    create_folder(project_name)
    #create_folder(os.path.join(project_name, "har"))
    create_folder(os.path.join(project_name, "testcases"))
    create_folder(os.path.join(project_name, "reports"))

    # create_file(
    #     os.path.join(project_name, "testcases", "demo_testcase_request.yml"),
    #     demo_testcase_request_content,
    # )
    # create_file(
    #     os.path.join(project_name, "testcases", "demo_testcase_ref.yml"),
    #     demo_testcase_with_ref_content,
    # )

    # 生成一个 Workbook 的实例化对象，wb即代表一个工作簿（一个 Excel 文件）
    wb = openpyxl.Workbook()
    # 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
    ws = wb.active
    # 更改工作表ws的title
    ws.title = 'test_sheet1'
    # 对ws的单个单元格传入数据
    ws['A1'] = 'case_id'
    ws['B1'] = 'name'
    ws['C1'] = 'setup_hooks'
    ws['D1'] = 'variables'
    ws['E1'] = 'contenttype'
    ws['F1'] = 'method'
    ws['G1'] = 'url'
    ws['H1'] = 'data'
    ws['I1'] = 'extract'
    ws['J1'] = 'validate'
    ws['K1'] = 'teardown_hooks'
    ws['L1'] = 'result'
    ws['M1'] = 'tag'

    ws['A2'] = '1'
    ws['B2'] = '添加事故记录'
    ws['C2'] = ''
    ws['D2'] = ''
    ws['E2'] = 'json'
    ws['F2'] = 'post'
    ws['G2'] = '/smartfactory/accidentRecord/add'
    ws['H2'] = '{"accidentLevel":"","accidentTime":"2021-06-10","accidentTypeId":1379597728769757186,"deptName":"部门A","fix":"无","fixStaffs":[{"createTime":"","createUserId":0,"id":0,"name":"肖利真","recordId":0,"userId":1374541488599445505}],"id":0,"implement":"落实情况A","lost":100,"name":"测试5295592321","reason":"无","relateStaffs":[{"createTime":"","createUserId":0,"id":0,"name":"陈安静","recordId":0,"userId":1374541488599445506}],"site":"位置A","type":1,"workGroup":"班组A"}'
    ws['I2'] = '{"recordId1": "body.result"}'
    ws['J2'] = '{"eq": ["body.innerCode", "200"]}'
    ws['K2'] = ''
    ws['L2'] = ''
    ws['M2'] = ''

    ws['A3'] = '2'
    ws['B3'] = '根据事故ID删除事故记录'
    ws['C3'] = ''
    ws['D3'] = ''
    ws['E3'] = 'data'
    ws['F3'] = 'post'
    ws['G3'] = '/smartfactory/accidentRecord/deleteById'
    ws['H3'] = '{"id":"$recordId1"}'
    ws['I3'] = ''
    ws['J3'] = '{"eq": ["body.innerCode", "200"]}'
    ws['K3'] = ''
    ws['L3'] = ''
    ws['M3'] = ''

    path = os.path.join(os.getcwd(), project_name, "testcases", "template.xlsx")
    wb.save(path)

    #写__init__文件
    if not os.path.exists(os.path.join(os.getcwd(), project_name, "testcases\\__init__.py")):
        f = open(os.path.join(os.getcwd(),project_name, "testcases\\__init__.py"), 'w')
        f.write(" ")
        f.close()


    #写conftest文件
    if not os.path.exists(os.path.join(os.getcwd(), project_name, "testcases\\conftest.py")):
        f = open(os.path.join(os.getcwd(), project_name, "testcases\\conftest.py"), 'w')
        f.write(content)
        f.close()

    #写config.ini文件
    if not os.path.exists(os.path.join(os.getcwd(), project_name, "config.ini")):
        f = open(os.path.join(os.getcwd(), project_name, "config.ini"), 'w', encoding='utf-8')
        f.write(config_content)
        f.close()


    create_file(os.path.join(project_name, "debugtalk.py"), demo_debugtalk_content)
    #create_file(os.path.join(project_name, ".env"), demo_env_content)
    create_file(os.path.join(project_name, ".gitignore"), ignore_content)

    show_tree(project_name)
    return 0


def main_scaffold(args):
    capture_message("startproject with scaffold")
    sys.exit(create_scaffold(args.project_name))
