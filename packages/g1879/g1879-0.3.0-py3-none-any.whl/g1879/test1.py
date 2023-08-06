#!/usr/bin/env python
# -*- coding:utf-8 -*-
import math

import openpyxl

# from g1879.xlsx import copy_cell, copy_row, 设置行高

e = r"C:\Users\g1879\Desktop\111\tmp3\3D第2轮 - 副本.xlsx"
# e2 = r"C:\Users\g1879\Desktop\111\tmp3\3D第2轮 - 副本.xlsx"
# import pandas as pd
# from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


wb = openpyxl.load_workbook(e)
ws = wb.active
col_width = []
for i in range(len(next(ws.iter_rows()))):
    col_letter = get_column_letter(i + 1)

    # minimum_width = 20
    # current_width = ws.column_dimensions[col_letter].width
    # if not current_width or current_width < minimum_width:
    #     ws.column_dimensions[col_letter].width = minimum_width

    col_width.append(ws.column_dimensions[col_letter].width)

for i, row in enumerate(ws):
    default_height = 12.5  # Corresponding to font size 12

    multiples_of_font_size = [default_height]
    for j, cell in enumerate(row):
        wrap_text = True
        vertical = "top"
        if cell.value is not None:
            mul = 0
            for v in str(cell.value).split('\n'):
                mul += math.ceil(len(v) / col_width[j]) * cell.font.size

            if mul > 0:
                multiples_of_font_size.append(mul)

        cell.alignment = Alignment(wrap_text=wrap_text, vertical=vertical)

    original_height = ws.row_dimensions[i + 1].height
    if original_height is None:
        original_height = default_height

    new_height = max(multiples_of_font_size)
    if original_height < new_height:
        ws.row_dimensions[i + 1].height = new_height
#
# # wb2 = openpyxl.Workbook()
# # ws2 = wb2.active
#
# # cell = ws['B3']
# # copy_row(ws,3,ws2,1)
# # copy_row()
# # e=copy_cell(cell, ws2, 3, 'A')
# # 设置行高(ws, 10)
# # 设置行高(ws, 16)
#
# # print(ws.column_dimensions['j'].width)
# ws.row_dimensions[1].bestFit = True
# # c=ws.cell(1,16)
# # ws.rows[1]
# # print(c.font)
# # print()
# # print(c.fill)
# # print()
# # print(c.number_format)
# # print()
# # print(c.protection)
# # print()
# # print(c.alignment)
# #
wb.save(e)
wb.close()
# # wb2.close()
